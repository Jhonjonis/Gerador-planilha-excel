import os
import sys
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from datetime import datetime
from tkinterdnd2 import TkinterDnD, DND_FILES

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import openpyxl.utils

# Lista para armazenar os cabeçalhos e os dados
cabecalhos = []
dados = []

# Verificação e instalação de pacotes
def check_and_install(package, import_name=None):
    try:
        __import__(import_name or package)
    except ImportError:
        print(f"O pacote {package} não está instalado.")
        user_input = input(f"Deseja instalar {package}? (S/N): ").strip().lower()
        if user_input == "s":
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        else:
            print(f"⚠️ O programa pode não funcionar corretamente sem {package}!")

required_packages = [("openpyxl", None), ("tkinterdnd2", "tkinterdnd2")]
for package, import_name in required_packages:
    check_and_install(package, import_name)

# Funções principais
def importar_planilha(arquivo=None):
    if not arquivo:
        arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
        if not arquivo:
            return
    else:
        arquivo = arquivo.strip("{}")

    try:
        wb = load_workbook(arquivo)
        sheet = wb.active

        cabecalhos.clear()
        dados.clear()

        for col in sheet.iter_cols(min_row=1, max_row=1, values_only=True):
            cabecalhos.append(col[0])

        for row in sheet.iter_rows(min_row=2, values_only=True):
            dados.append(list(row))

        messagebox.showinfo("Sucesso", "Planilha importada com sucesso!")
        atualizar_entrada_de_dados()
        atualizar_lista_dados()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao importar planilha: {e}")

def definir_cabecalhos():
    global cabecalhos
    cabecalhos.clear()
    num_campos = simpledialog.askinteger("Cabeçalhos", "Quantos campos deseja adicionar?", minvalue=1)

    if num_campos:
        for i in range(num_campos):
            nome_campo = simpledialog.askstring("Cabeçalho", f"Digite o nome do cabeçalho {i + 1}:")
            if nome_campo:
                cabecalhos.append(nome_campo)

        messagebox.showinfo("Sucesso", f"Cabeçalhos definidos: {', '.join(cabecalhos)}")
        atualizar_entrada_de_dados()

def atualizar_entrada_de_dados():
    for widget in frame_entradas.winfo_children():
        widget.destroy()

    global entry_fields
    entry_fields = []

    max_columns = 2
    current_column = 0
    current_row = 0

    for i, cabecalho in enumerate(cabecalhos):
        ttk.Label(frame_entradas, text=cabecalho, font=("Arial", 11, "bold")).grid(
            row=current_row, column=current_column * 2, padx=5, pady=5, sticky="w"
        )

        entry = ttk.Entry(frame_entradas, font=("Arial", 11), width=30)
        entry.grid(row=current_row, column=(current_column * 2) + 1, padx=5, pady=5, sticky="ew")
        entry_fields.append(entry)

        current_column += 1
        if current_column >= max_columns:
            current_column = 0
            current_row += 1

    for i in range(max_columns * 2):
        frame_entradas.grid_columnconfigure(i, weight=1)

    frame_entradas.grid_rowconfigure(current_row, weight=1)

    lista_dados["columns"] = cabecalhos
    for col in cabecalhos:
        lista_dados.heading(col, text=col)
        lista_dados.column(col, anchor="center", width=150)

def atualizar_lista_dados():
    lista_dados.delete(*lista_dados.get_children())
    for row in dados:
        lista_dados.insert("", "end", values=row)

def adicionar_dados():
    if not cabecalhos:
        messagebox.showwarning("Aviso", "Defina os cabeçalhos antes de adicionar dados.")
        return

    if all(not entry.get().strip() for entry in entry_fields):
        messagebox.showwarning("Aviso", "Preencha pelo menos um campo antes de adicionar.")
        return

    valores = [entry.get() for entry in entry_fields]
    dados.append(valores)
    lista_dados.insert("", "end", values=valores)

    for entry in entry_fields:
        entry.delete(0, tk.END)

def editar_dado():
    try:
        item_selecionado = lista_dados.selection()[0]
        valores_atuais = lista_dados.item(item_selecionado, "values")

        if not valores_atuais:
            return

        janela_edicao = tk.Toplevel(root)
        janela_edicao.title("Editar Dados")

        entradas_edicao = []
        for i, valor in enumerate(valores_atuais):
            ttk.Label(janela_edicao, text=cabecalhos[i]).grid(row=i, column=0, padx=5, pady=5, sticky="w")
            entrada = ttk.Entry(janela_edicao, font=("Arial", 11))
            entrada.insert(0, valor)
            entrada.grid(row=i, column=1, padx=5, pady=5)
            entradas_edicao.append(entrada)

        def salvar_edicao():
            novos_valores = [entrada.get() for entrada in entradas_edicao]
            lista_dados.item(item_selecionado, values=novos_valores)
            index = lista_dados.index(item_selecionado)
            dados[index] = novos_valores
            janela_edicao.destroy()

        ttk.Button(janela_edicao, text="Salvar", command=salvar_edicao).grid(columnspan=2, pady=10)

    except IndexError:
        messagebox.showwarning("Aviso", "Selecione um item para editar.")

def salvar_em_excel():
    if not cabecalhos:
        messagebox.showerror("Erro", "Defina os cabeçalhos antes de salvar.")
        return

    if not dados:
        messagebox.showwarning("Aviso", "Nenhum dado para salvar.")
        return

    pasta_planilhas = "planilhas"
    if not os.path.exists(pasta_planilhas):
        os.makedirs(pasta_planilhas)

    nome_arquivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Arquivos Excel", "*.xlsx")],
                                                 initialdir=pasta_planilhas,
                                                 initialfile=f"dados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    if not nome_arquivo:
        return

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Dados"

    bold_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for col_num, header in enumerate(cabecalhos, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    for row_num, row_data in enumerate(dados, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border

    for i, _ in enumerate(cabecalhos, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    wb.save(nome_arquivo)
    messagebox.showinfo("Sucesso", f"Dados salvos em {nome_arquivo}")

    dados.clear()
    lista_dados.delete(*lista_dados.get_children())

# Interface gráfica principal
root = TkinterDnD.Tk()
root.title("Gerador de Planilhas")
root.geometry("700x700")
root.configure(bg="#f0f0f0")

style = ttk.Style()
style.configure("TButton", font=("Arial", 11), padding=5)
style.configure("TLabel", background="#f0f0f0", font=("Arial", 11))
style.configure("Treeview.Heading", font=("Arial", 11, "bold"))

frame_principal = ttk.Frame(root, padding=10)
frame_principal.pack(fill="both", expand=True)

btn_definir_cabecalhos = ttk.Button(frame_principal, text="Definir Cabeçalhos", command=definir_cabecalhos)
btn_definir_cabecalhos.pack(pady=10)

frame_entradas = ttk.LabelFrame(frame_principal, text="Entradas", padding=10)
frame_entradas.pack(fill="both", expand=True, padx=10, pady=10)

btn_adicionar = ttk.Button(frame_principal, text="Adicionar Dados", command=adicionar_dados)
btn_adicionar.pack(pady=10)

frame_lista = ttk.LabelFrame(frame_principal, text="Dados Adicionados", padding=10)
frame_lista.pack(fill="both", expand=True, padx=10, pady=10)

lista_dados = ttk.Treeview(frame_lista, show="headings", height=8)
lista_dados.pack(fill="both", expand=True)

lista_dados.bind("<Double-1>", lambda event: editar_dado())

frame_botoes_final = ttk.Frame(frame_principal)
frame_botoes_final.pack(pady=10, fill="x", side="bottom")

btn_gerar_planilha = ttk.Button(frame_botoes_final, text="Gerar Planilha", command=salvar_em_excel)
btn_gerar_planilha.pack(side="left", padx=10)

btn_importar_planilha = ttk.Button(frame_botoes_final, text="Importar Planilha", command=importar_planilha)
btn_importar_planilha.pack(side="left", padx=10)

# Suporte a arrastar e soltar
root.drop_target_register(DND_FILES)
root.dnd_bind('<<Drop>>', lambda event: importar_planilha(event.data))

if __name__ == "__main__":
    root.mainloop()
